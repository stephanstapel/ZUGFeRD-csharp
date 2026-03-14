#!/usr/bin/env python3
"""
ZUGFeRD 1.0 Schema Extractor
=============================
Parses the XSD schema files and Schematron (SCMT) file from
documentation/zugferd10/Schema/ and extracts:
  - XML element structure (names, types, parent-child relationships)
  - Cardinality (minOccurs / maxOccurs from XSD)
  - Profile-specific cardinality rules from Schematron (BASIC / COMFORT / EXTENDED)

BT/BG numbers and textual descriptions are NOT present in the ZUGFeRD 1.0
documentation files — they were introduced with the EN16931 standard (ZUGFeRD
2.x / Factur-X).  This script cross-references the element names against the
ZUGFeRD 2.x EN16931 Schematron file shipped with this repository to provide
best-effort BT/BG number annotations and English descriptions.

Outputs:
  - documentation/zugferd10/schema_structure.json
  - documentation/zugferd10/schema_structure.xlsx

Usage:
    python3 tools/extract_zugferd10_schema.py
    python3 tools/extract_zugferd10_schema.py --schema-dir path/to/Schema
                                               --output-dir path/to/output
"""

import argparse
import json
import re
import sys
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel output will be skipped.", file=sys.stderr)
    print("Install with: pip install openpyxl", file=sys.stderr)


# ---------------------------------------------------------------------------
# Namespace constants
# ---------------------------------------------------------------------------
NS_XS = "http://www.w3.org/2001/XMLSchema"
NS_RSM = "urn:ferd:CrossIndustryDocument:invoice:1p0"
NS_RAM = "urn:un:unece:uncefact:data:standard:ReusableAggregateBusinessInformationEntity:12"
NS_QDT = "urn:un:unece:uncefact:data:standard:QualifiedDataType:12"
NS_UDT = "urn:un:unece:uncefact:data:standard:UnqualifiedDataType:15"

NS_MAP = {
    "rsm": NS_RSM,
    "ram": NS_RAM,
    "qdt": NS_QDT,
    "udt": NS_UDT,
    "xs":  NS_XS,
}

# Reverse lookup: namespace URI → short prefix
NS_PREFIX = {v: k for k, v in NS_MAP.items()}

XSD_FILES = [
    "ZUGFeRD1p0.xsd",
    "ZUGFeRD1p0_urn_un_unece_uncefact_data_standard_ReusableAggregateBusinessInformationEntity_12.xsd",
    "ZUGFeRD1p0_urn_un_unece_uncefact_data_standard_QualifiedDataType_12.xsd",
    "ZUGFeRD1p0_urn_un_unece_uncefact_data_standard_UnqualifiedDataType_15.xsd",
]

SCMT_FILE = "ZUGFeRD_1p0.scmt"

# Path to the ZUGFeRD 2.x EN16931 Schematron file used for BT/BG number enrichment.
# This path is relative to the repository root and resolved at runtime.
EN16931_SCH_REL = (
    "documentation/zugferd211en/Schema/EN16931/Schematron/FACTUR-X_EN16931.sch"
)


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def clark_to_prefix(clark_name: str) -> str:
    """Convert Clark notation {uri}local to prefix:local."""
    m = re.match(r"^\{([^}]+)\}(.+)$", clark_name)
    if m:
        uri, local = m.group(1), m.group(2)
        prefix = NS_PREFIX.get(uri, uri)
        return f"{prefix}:{local}"
    return clark_name


def strip_ns(clark_name: str) -> str:
    """Return local part of Clark notation."""
    m = re.match(r"^\{[^}]+\}(.+)$", clark_name)
    return m.group(1) if m else clark_name


def type_ref_to_str(type_ref: str | None) -> str:
    """Normalise a type reference (Clark or prefix:local) to prefix:local."""
    if not type_ref:
        return ""
    return clark_to_prefix(type_ref) if type_ref.startswith("{") else type_ref


def prefix_of_type(type_name: str) -> str:
    """Return the namespace prefix of a qualified type name like 'ram:TradePartyType'."""
    if ":" in type_name:
        return type_name.split(":")[0]
    return ""


# ---------------------------------------------------------------------------
# XSD Parser
# ---------------------------------------------------------------------------

class XsdParser:
    """
    Parses a set of XSD files and builds a map of type → element list.

    Each entry in the map has the form:
        type_name (str, e.g. 'ram:TradePartyType') → list of ElementInfo dicts

    Each ElementInfo dict contains:
        name        – local element name
        type        – qualified type reference (e.g. 'udt:IDType')
        minOccurs   – string value from XSD ('0', '1', ...)
        maxOccurs   – string value from XSD ('0', '1', 'unbounded', ...)
        nsPrefix    – namespace prefix of the *enclosing* type (inherited by elements
                      defined in that XSD when elementFormDefault="qualified")
    """

    def __init__(self) -> None:
        # Map from fully qualified type name (prefix:local) to list of element dicts
        self.types: dict[str, list[dict[str, Any]]] = {}
        # Simple type names
        self.simple_types: set[str] = set()
        # Root elements (local name → qualified type ref)
        self.root_elements: dict[str, str] = {}
        # Map type name → namespace prefix (e.g. 'rsm:CrossIndustryDocumentType' → 'rsm')
        self.type_ns: dict[str, str] = {}

    def parse_file(self, filepath: str) -> None:
        tree = ET.parse(filepath)
        root = tree.getroot()

        # Determine target namespace
        target_ns = root.get("targetNamespace", "")
        prefix = NS_PREFIX.get(target_ns, "tns") if target_ns else ""

        self._process_schema(root, prefix)

    def _qualify(self, local: str, prefix: str) -> str:
        return f"{prefix}:{local}" if prefix else local

    def _process_schema(self, schema_root: ET.Element, prefix: str) -> None:
        for child in schema_root:
            tag = strip_ns(child.tag)
            name = child.get("name", "")

            if tag == "element":
                type_ref = child.get("type", "")
                self.root_elements[name] = type_ref_to_str(type_ref)

            elif tag == "complexType":
                qname = self._qualify(name, prefix)
                elements = self._parse_complex_type(child, prefix)
                self.types[qname] = elements
                self.type_ns[qname] = prefix

            elif tag == "simpleType":
                qname = self._qualify(name, prefix)
                self.simple_types.add(qname)

    def _parse_complex_type(
        self, ct_elem: ET.Element, enclosing_prefix: str
    ) -> list[dict[str, Any]]:
        """Extract all elements from a complexType definition."""
        elements: list[dict[str, Any]] = []

        for child in ct_elem:
            tag = strip_ns(child.tag)
            if tag in ("sequence", "choice", "all"):
                elements.extend(self._parse_group(child, enclosing_prefix))
            elif tag in ("simpleContent", "complexContent"):
                for grandchild in child:
                    if strip_ns(grandchild.tag) == "extension":
                        elements.extend(self._parse_group(grandchild, enclosing_prefix))

        return elements

    def _parse_group(
        self, group_elem: ET.Element, enclosing_prefix: str
    ) -> list[dict[str, Any]]:
        """Recursively parse xs:sequence / xs:choice / xs:all containers."""
        elements: list[dict[str, Any]] = []
        for child in group_elem:
            tag = strip_ns(child.tag)
            if tag == "element":
                elem_name = child.get("name", "")
                type_ref = type_ref_to_str(child.get("type", ""))
                min_occurs = child.get("minOccurs", "1")
                max_occurs = child.get("maxOccurs", "1")
                elem: dict[str, Any] = {
                    "name": elem_name,
                    "type": type_ref,
                    "minOccurs": min_occurs,
                    "maxOccurs": max_occurs,
                    # Elements defined inside an xs:complexType belonging to
                    # a given namespace are qualified with that namespace.
                    "nsPrefix": enclosing_prefix,
                }
                # Inline anonymous complex type
                for subchild in child:
                    if strip_ns(subchild.tag) == "complexType":
                        elem["inlineType"] = self._parse_complex_type(subchild, enclosing_prefix)
                elements.append(elem)
            elif tag in ("sequence", "choice", "all"):
                elements.extend(self._parse_group(child, enclosing_prefix))

        return elements


# ---------------------------------------------------------------------------
# Schematron (SCMT) Parser
# ---------------------------------------------------------------------------

class ScmtParser:
    """
    Parses a Schematron (.scmt) file and extracts profile-specific cardinality
    rules for ZUGFeRD 1.0 (BASIC, COMFORT, EXTENDED).

    Each pattern/rule block maps an XPath context to a set of constraints.
    """

    SCH_NS = "http://purl.oclc.org/dsdl/schematron"

    def __init__(self) -> None:
        self.rules: list[dict[str, Any]] = []

    def parse_file(self, filepath: str) -> None:
        tree = ET.parse(filepath)
        root = tree.getroot()

        for pattern in root.iter(f"{{{self.SCH_NS}}}pattern"):
            for rule in pattern.iter(f"{{{self.SCH_NS}}}rule"):
                context = rule.get("context", "")
                constraints: list[dict[str, str]] = []

                for child in rule:
                    child_tag = strip_ns(child.tag)
                    if child_tag == "assert":
                        test = child.get("test", "")
                        text = (child.text or "").strip()
                        constraints.append(
                            {"type": "assert", "test": test, "message": text}
                        )
                    elif child_tag == "report":
                        test = child.get("test", "")
                        text = (child.text or "").strip()
                        constraints.append(
                            {"type": "report", "test": test, "message": text}
                        )

                if constraints:
                    self.rules.append({"context": context, "constraints": constraints})

    def get_context_map(self) -> dict[str, list[dict[str, str]]]:
        """Return a dict mapping XPath context → list of constraints."""
        result: dict[str, list[dict[str, str]]] = {}
        for rule in self.rules:
            ctx = rule["context"]
            result.setdefault(ctx, []).extend(rule["constraints"])
        return result


# ---------------------------------------------------------------------------
# BT/BG Number Parser (ZUGFeRD 2.x EN16931 Schematron)
# ---------------------------------------------------------------------------

class BtMappingParser:
    """
    Parses a ZUGFeRD 2.x / Factur-X EN16931 Schematron (.sch) file and
    extracts BT (Business Term) and BG (Business Group) number associations.

    ZUGFeRD 1.0 does not contain BT numbers — they were introduced with the
    EN16931 standard.  However, because both ZUGFeRD 1.0 and 2.x are based on
    the UN/CEFACT Cross Industry Invoice (CII) model, many element *local names*
    in the ``ram:`` namespace are identical across versions.  This class builds a
    best-effort mapping from element local names (and their parent context) to
    BT/BG numbers so that the ZUGFeRD 1.0 schema export can include them.

    Mapping key:
        (parent_context_tail_local_name, element_local_name)  →  BtInfo

    A fallback mapping keyed only on element local name is also provided for
    elements that appear in a single context.
    """

    SCH_NS = "http://purl.oclc.org/dsdl/schematron"
    BT_RE = re.compile(r"\(?(B[TG]-\d+)\)?")
    # Matches a simple element reference in an assert test, e.g. "(ram:BasisAmount)"
    SIMPLE_ELEM_RE = re.compile(r"^\s*\(?\s*([a-z]+:[A-Za-z0-9]+)\s*\)?\s*$")

    def __init__(self) -> None:
        # (parent_local, elem_local) → {"bts": [...], "description": "..."}
        self._ctx_map: dict[tuple[str, str], dict[str, Any]] = {}
        # elem_local → {"bts": [...], "description": "..."}   (fallback, single-context elems)
        self._elem_map: dict[str, dict[str, Any]] = {}

    def parse_file(self, filepath: str) -> None:
        tree = ET.parse(filepath)
        root = tree.getroot()

        for pattern in root.iter(f"{{{self.SCH_NS}}}pattern"):
            for rule in pattern.iter(f"{{{self.SCH_NS}}}rule"):
                context = rule.get("context", "")
                # Extract the local name of the innermost non-predicate element
                ctx_local = self._context_tail(context)

                for a in rule.iter(f"{{{self.SCH_NS}}}assert"):
                    test = a.get("test", "")
                    msg = re.sub(r"\[[A-Z0-9-]+\]-", "", (a.text or "")).strip()
                    bts = self.BT_RE.findall(msg)
                    if not bts:
                        continue

                    m = self.SIMPLE_ELEM_RE.match(test)
                    if not m:
                        continue
                    elem_local = m.group(1).split(":")[-1]

                    key = (ctx_local, elem_local)
                    entry = self._ctx_map.setdefault(key, {"bts": [], "description": ""})
                    for bt in bts:
                        if bt not in entry["bts"]:
                            entry["bts"].append(bt)
                    if not entry["description"] and msg:
                        entry["description"] = msg[:200]

        # Build the fallback elem_map for elements that map to exactly one BT
        elem_bt_count: dict[str, set[str]] = {}
        elem_desc: dict[str, str] = {}
        for (ctx_local, elem_local), entry in self._ctx_map.items():
            elem_bt_count.setdefault(elem_local, set()).update(entry["bts"])
            elem_desc.setdefault(elem_local, entry["description"])

        for elem_local, bts in elem_bt_count.items():
            self._elem_map[elem_local] = {
                "bts": sorted(bts),
                "description": elem_desc.get(elem_local, ""),
            }

    def lookup(
        self, parent_local: str, elem_local: str
    ) -> dict[str, Any]:
        """
        Return the best BT mapping for (parent_local, elem_local).

        Priority:
        1. Exact (parent, elem) key match
        2. Fallback elem-only match
        3. Empty result
        """
        exact = self._ctx_map.get((parent_local, elem_local))
        if exact:
            return exact
        return self._elem_map.get(elem_local, {"bts": [], "description": ""})

    # ------------------------------------------------------------------
    @staticmethod
    def _context_tail(context: str) -> str:
        """
        Return the local name of the innermost element in an XPath context
        expression, stripping predicates.

        E.g. ``//ram:ApplicableHeaderTradeSettlement/ram:ApplicableTradeTax``
        → ``"ApplicableTradeTax"``
        """
        parts = [p for p in context.split("/") if p and ":" in p]
        if not parts:
            return ""
        return parts[-1].split("[")[0].split(":")[-1]


# ---------------------------------------------------------------------------
# Tree Builder
# ---------------------------------------------------------------------------

def build_tree(
    xsd: XsdParser,
    scmt_map: dict[str, list[dict[str, str]]],
    type_name: str,
    element_name: str,
    xpath: str,
    bt_map: "BtMappingParser | None" = None,
    parent_local: str = "",
    visited: set[str] | None = None,
) -> dict[str, Any]:
    """
    Recursively build an element tree node starting from type_name.

    XPaths are built with namespace prefixes matching the Schematron file so
    that Schematron constraints can be correlated to each node.

    bt_map, when provided, enriches each node with BT/BG number annotations
    and EN16931 descriptions cross-referenced from the ZUGFeRD 2.x Schematron.

    visited prevents infinite recursion for circular type references.
    """
    if visited is None:
        visited = set()

    constraints = scmt_map.get(xpath, [])

    bt_info: dict[str, Any] = {"bts": [], "description": ""}
    if bt_map is not None:
        bt_info = bt_map.lookup(parent_local, element_name)

    node: dict[str, Any] = {
        "name": element_name,
        "type": type_name,
        "xpath": xpath,
        "bt_numbers": bt_info["bts"],
        "bt_description": bt_info["description"],
        "constraints": constraints,
        "children": [],
    }

    if type_name in visited:
        node["note"] = "circular reference – not expanded"
        return node

    if type_name not in xsd.types:
        return node

    # Derive the local name of the current type to use as the parent context
    # for BT lookups of child elements.
    type_local = type_name.split(":")[-1] if ":" in type_name else type_name
    # Strip the "Type" suffix to get the element context name used in Schematron
    # e.g. "ram:ApplicableTradeTaxType" → "ApplicableTradeTax"
    bt_parent_local = type_local.removesuffix("Type")

    visited = visited | {type_name}
    for child_elem in xsd.types[type_name]:
        child_name = child_elem["name"]
        child_type = child_elem["type"]
        # The XPath segment prefix comes from the namespace of the *parent type*
        # (elementFormDefault="qualified" means local elements take the target namespace
        # of the schema file that defines the enclosing complexType).
        child_ns_prefix = child_elem.get("nsPrefix", "")
        if child_ns_prefix:
            child_xpath = f"{xpath}/{child_ns_prefix}:{child_name}"
        else:
            child_xpath = f"{xpath}/{child_name}"

        child_bt_info: dict[str, Any] = {"bts": [], "description": ""}
        if bt_map is not None:
            child_bt_info = bt_map.lookup(bt_parent_local, child_name)

        child_node: dict[str, Any] = {
            "name": child_name,
            "type": child_type,
            "minOccurs": child_elem["minOccurs"],
            "maxOccurs": child_elem["maxOccurs"],
            "xpath": child_xpath,
            "bt_numbers": child_bt_info["bts"],
            "bt_description": child_bt_info["description"],
            "constraints": scmt_map.get(child_xpath, []),
            "children": [],
        }
        if "inlineType" in child_elem:
            child_node["inlineElements"] = child_elem["inlineType"]

        # Recurse into child type
        if child_type and child_type in xsd.types:
            subtree = build_tree(
                xsd, scmt_map, child_type, child_name, child_xpath,
                bt_map, bt_parent_local, visited
            )
            child_node["children"] = subtree["children"]

        node["children"].append(child_node)

    return node


# ---------------------------------------------------------------------------
# Flat row generator (for Excel / tabular output)
# ---------------------------------------------------------------------------

def flatten_tree(
    node: dict[str, Any],
    rows: list[dict[str, Any]] | None = None,
    depth: int = 0,
) -> list[dict[str, Any]]:
    """Convert the hierarchical tree into a flat list of rows."""
    if rows is None:
        rows = []

    min_occ = node.get("minOccurs", "")
    max_occ = node.get("maxOccurs", "")

    assert_msgs = [
        c["message"] for c in node.get("constraints", []) if c.get("type") == "assert"
    ]
    report_msgs = [
        c["message"] for c in node.get("constraints", []) if c.get("type") == "report"
    ]

    rows.append(
        {
            "depth": depth,
            "name": node.get("name", ""),
            "type": node.get("type", ""),
            "xpath": node.get("xpath", ""),
            "minOccurs": min_occ,
            "maxOccurs": max_occ,
            "schematron_required": "; ".join(assert_msgs),
            "schematron_not_used": "; ".join(report_msgs),
        }
    )

    for child in node.get("children", []):
        flatten_tree(child, rows, depth + 1)

    return rows


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

HEADER_FILL = (
    PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    if OPENPYXL_AVAILABLE
    else None
)
HEADER_FONT = Font(bold=True, color="FFFFFF") if OPENPYXL_AVAILABLE else None
DEPTH_COLORS = [
    "D6E4F0",
    "AED6F1",
    "85C1E9",
    "5DADE2",
    "3498DB",
    "2E86C1",
]


def write_excel(rows: list[dict[str, Any]], output_path: str) -> None:
    if not OPENPYXL_AVAILABLE:
        print("Excel output skipped (openpyxl not installed).")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ZUGFeRD 1.0 Schema"

    columns = [
        ("XPath", 60),
        ("Element Name", 40),
        ("Data Type", 45),
        ("minOccurs", 10),
        ("maxOccurs", 10),
        ("Depth", 8),
        ("Schematron – Required (assert)", 50),
        ("Schematron – Not Used (report)", 50),
    ]

    # Write header row
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 25

    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        depth = row["depth"]
        indent = "  " * depth
        fill_color = DEPTH_COLORS[min(depth, len(DEPTH_COLORS) - 1)]
        row_fill = PatternFill(
            start_color=fill_color, end_color=fill_color, fill_type="solid"
        )

        values = [
            row["xpath"],
            indent + row["name"],
            row["type"],
            row["minOccurs"],
            row["maxOccurs"],
            depth,
            row["schematron_required"],
            row["schematron_not_used"],
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"Excel written: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    repo_root = script_dir.parent
    default_schema = str(repo_root / "documentation" / "zugferd10" / "Schema")
    default_output = str(repo_root / "documentation" / "zugferd10")

    parser = argparse.ArgumentParser(
        description="Extract ZUGFeRD 1.0 XML schema structure to JSON and Excel."
    )
    parser.add_argument(
        "--schema-dir",
        default=default_schema,
        help=f"Path to Schema directory (default: {default_schema})",
    )
    parser.add_argument(
        "--output-dir",
        default=default_output,
        help=f"Path for output files (default: {default_output})",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Skip Excel output even if openpyxl is available",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    schema_dir = Path(args.schema_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # ---- Parse XSD files ----
    xsd = XsdParser()
    for xsd_file in XSD_FILES:
        path = schema_dir / xsd_file
        if not path.exists():
            print(f"Warning: XSD file not found: {path}", file=sys.stderr)
            continue
        print(f"Parsing XSD: {path.name}")
        xsd.parse_file(str(path))

    print(
        f"Loaded {len(xsd.types)} complex types, "
        f"{len(xsd.simple_types)} simple types, "
        f"{len(xsd.root_elements)} root elements."
    )

    # ---- Parse Schematron ----
    scmt_path = schema_dir / SCMT_FILE
    scmt_map: dict[str, list[dict[str, str]]] = {}
    if scmt_path.exists():
        print(f"Parsing Schematron: {scmt_path.name}")
        scmt = ScmtParser()
        scmt.parse_file(str(scmt_path))
        scmt_map = scmt.get_context_map()
        print(f"Loaded {len(scmt_map)} Schematron rule contexts.")
    else:
        print(f"Warning: Schematron file not found: {scmt_path}", file=sys.stderr)

    # ---- Build element tree ----
    # The root element is CrossIndustryDocument (rsm namespace)
    root_elem_name = "CrossIndustryDocument"
    root_type = xsd.root_elements.get(root_elem_name, "rsm:CrossIndustryDocumentType")
    root_xpath = f"/rsm:{root_elem_name}"

    print(f"Building element tree from root: {root_elem_name} ({root_type})")
    tree = build_tree(xsd, scmt_map, root_type, root_elem_name, root_xpath)

    # ---- JSON output ----
    json_path = output_dir / "schema_structure.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(tree, f, indent=2, ensure_ascii=False)
    print(f"JSON written:  {json_path}")

    # ---- Excel output ----
    if not args.no_excel:
        rows = flatten_tree(tree)
        print(f"Total element rows: {len(rows)}")
        xlsx_path = output_dir / "schema_structure.xlsx"
        write_excel(rows, str(xlsx_path))
    else:
        print("Excel output skipped (--no-excel).")

    print("Done.")


if __name__ == "__main__":
    main()
